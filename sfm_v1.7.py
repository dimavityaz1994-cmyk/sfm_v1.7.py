import os
import pandas as pd
import xml.etree.ElementTree as ET
from datetime import datetime
import customtkinter as ctk
from tkinter import filedialog, messagebox
import tkinter.ttk as ttk
import requests
import threading
from collections import Counter
from urllib.parse import quote

# Попытка импорта openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ════════════════════════════════════════════════════════════════════════════
#  ПУТИ И URL
# ════════════════════════════════════════════════════════════════════════════

BASE_EXCEL_PATH = r"K:\COMPLIANCE\AML\Террористы в Сравни!.xlsx"
MFO_LOCAL_PATH  = r"K:\COMPLIANCE\AML\Мониторинг\Проверки МФО\МФО на обслуживании.xlsx"
MFO_CBR_URL     = "https://www.cbr.ru/vfs/finmarkets/files/supervision/list_MFO.xlsx"

# Динамическая ссылка банков — дата подставляется при запросе
def get_banks_cbr_url():
    today = datetime.now().strftime("%m/%d/%Y")  # MM/DD/YYYY
    date_enc = today.replace("/", "%2F")
    return (
        f"https://www.cbr.ru/Queries/UniDbQuery/DownloadExcel/98547"
        f"?FromDate={date_enc}&ToDate={date_enc}"
        f"&posted=False&backUrl=%2Fbanking_sector%2Fcredit%2FFullCoList%2F"
    )

# ════════════════════════════════════════════════════════════════════════════
#  УТИЛИТЫ
# ════════════════════════════════════════════════════════════════════════════

def normalize(text):
    return " ".join(str(text).upper().replace("Ё", "Е").split())

def format_date(value):
    try: return pd.to_datetime(value).strftime("%Y-%m-%d")
    except: return ""

def format_date_ru(value):
    try: return pd.to_datetime(value).strftime("%d.%m.%Y")
    except: return ""

def parse_xml_date(date_str):
    try: return datetime.strptime(date_str, "%Y-%m-%d").date()
    except: return None

def clean_ogrn(value):
    """ОГРН может начинаться с нуля — сохраняем как строку."""
    if pd.isna(value):
        return ""
    s = str(value).strip()
    # Убираем .0 если число было прочитано как float
    if s.endswith(".0"):
        s = s[:-2]
    # Убираем пробелы
    s = s.replace(" ", "")
    return s

# ════════════════════════════════════════════════════════════════════════════
#  GUI — ТЕМА
# ════════════════════════════════════════════════════════════════════════════

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# ── Slate Pastel · Вариант B (Синий туман) ──────────────────────────────────
CLR_BG        = "#dce3ec"   # фон — slate pastel
CLR_SURFACE   = "#eef2f7"   # поверхность карточек
CLR_SURFACE2  = "#e2eaf3"   # поверхность вторичная
CLR_BORDER    = "#c4d4e4"   # рамки — голубоватые
CLR_ACCENT    = "#5a94c0"   # акцент — синий туман
CLR_ACCENT2   = "#4a80aa"   # акцент hover
CLR_SUCCESS   = "#4aaa6a"   # зелёный
CLR_DANGER    = "#c05a5a"   # красный приглушённый
CLR_WARNING   = "#b08030"   # жёлтый приглушённый
CLR_PURPLE    = "#8070b0"   # фиолетовый приглушённый
CLR_TEXT      = "#2a3a46"   # основной текст тёмный
CLR_MUTED     = "#6a8090"   # приглушённый текст
CLR_HIGHLIGHT = "#d0e4f4"   # подсветка строки

root = ctk.CTk()
root.title("СФМ v1.7")
root.geometry("560x480")
root.configure(fg_color=CLR_BG)

current_frame = None
def clear_frame():
    global current_frame
    if current_frame:
        current_frame.destroy()

# ── Treeview стиль ──────────────────────────────────────────────────────────
style = ttk.Style()
style.theme_use("clam")
style.configure("Treeview",
                background=CLR_SURFACE, foreground=CLR_TEXT,
                fieldbackground=CLR_SURFACE, bordercolor=CLR_BORDER,
                borderwidth=0, font=("Bahnschrift", 13), rowheight=32)
style.configure("Treeview.Heading",
                background=CLR_SURFACE2, foreground=CLR_MUTED,
                bordercolor=CLR_BORDER, borderwidth=0,
                font=("Bahnschrift", 13, "bold"), relief="flat")
style.map("Treeview",
          background=[("selected", CLR_ACCENT)],
          foreground=[("selected", "#ffffff")])
style.map("Treeview.Heading",
          background=[("active", CLR_BORDER)])

# ════════════════════════════════════════════════════════════════════════════
#  АНИМАЦИИ
# ════════════════════════════════════════════════════════════════════════════

class LoadingOverlay:
    SPINNER = ["▱▱▱▱▱", "▰▱▱▱▱", "▰▰▱▱▱", "▰▰▰▱▱", "▰▰▰▰▱", "▰▰▰▰▰", "▱▰▰▰▰", "▱▱▰▰▰", "▱▱▱▰▰", "▱▱▱▱▰"]

    def __init__(self, parent):
        self._running  = False
        self._spin_idx = 0
        self.overlay = ctk.CTkFrame(parent, fg_color="#dce3ec", corner_radius=0)
        self.card    = ctk.CTkFrame(self.overlay, fg_color=CLR_SURFACE,
                                    corner_radius=16, border_width=1, border_color=CLR_BORDER)
        self.card.place(relx=0.5, rely=0.5, anchor="center")

        self.spinner_lbl = ctk.CTkLabel(self.card, text="⠋",
                                        font=("Bahnschrift", 32), text_color=CLR_ACCENT)
        self.spinner_lbl.pack(pady=(28, 4))
        self.status_lbl = ctk.CTkLabel(self.card, text="Загрузка...",
                                       font=("Bahnschrift", 14, "bold"), text_color=CLR_TEXT)
        self.status_lbl.pack(pady=(0, 4))
        self.sub_lbl = ctk.CTkLabel(self.card, text="",
                            font=("Bahnschrift", 11), text_color=CLR_MUTED,
                            wraplength=320, justify="left")
        self.sub_lbl.pack(pady=(0, 14))
        self.progress_bar = ctk.CTkProgressBar(self.card, width=300, height=8,
                                               corner_radius=4, fg_color=CLR_SURFACE2,
                                               progress_color=CLR_ACCENT)
        self.progress_bar.set(0)
        self.progress_bar.pack(padx=28, pady=(0, 6))
        self.pct_lbl = ctk.CTkLabel(self.card, text="0%",
                                    font=("Bahnschrift", 11), text_color=CLR_MUTED)
        self.pct_lbl.pack(pady=(2, 24))

    def show(self, status="Загрузка...", sub=""):
        self.overlay.place(relx=0, rely=0, relwidth=1, relheight=1)
        self.overlay.lift()
        self.status_lbl.configure(text=status)
        self.sub_lbl.configure(text=sub)
        self.progress_bar.set(0)
        self.pct_lbl.configure(text="0%")
        self._running = True
        self._animate()

    def hide(self):
        self._running = False
        self.overlay.place_forget()

    def set_progress(self, value, status=None, sub=None):
        v = max(0.0, min(1.0, value))
        self.progress_bar.set(v)
        self.pct_lbl.configure(text=f"{int(v * 100)}%")
        if status:       self.status_lbl.configure(text=status)
        if sub is not None: self.sub_lbl.configure(text=sub)

    def _animate(self):
        if not self._running: return
        self.spinner_lbl.configure(text=self.SPINNER[self._spin_idx % len(self.SPINNER)])
        self._spin_idx += 1
        self.overlay.after(120, self._animate)


class ToastNotification:
    def __init__(self, parent):
        self._visible = False
        self.frame = ctk.CTkFrame(parent, fg_color=CLR_SURFACE2,
                                  corner_radius=10, border_width=1, border_color=CLR_BORDER)
        self.icon_lbl = ctk.CTkLabel(self.frame, text="✅", font=("Bahnschrift", 16))
        self.icon_lbl.pack(side="left", padx=(14, 6), pady=10)
        self.text_lbl = ctk.CTkLabel(self.frame, text="",
                                     font=("Bahnschrift", 13), text_color=CLR_TEXT)
        self.text_lbl.pack(side="left", padx=(0, 20), pady=10)

    def show(self, message, icon="✅", duration=3000):
        self.icon_lbl.configure(text=icon)
        self.text_lbl.configure(text=message)
        self.frame.place(relx=0.5, rely=0.96, anchor="s")
        self.frame.lift()
        self._visible = True
        self.frame.after(duration, self.hide)

    def hide(self):
        if self._visible:
            self.frame.place_forget()
            self._visible = False


def animate_rows(tree, rows, tags_list, delay=18):
    def _insert(idx):
        if idx >= len(rows): return
        tag = tags_list[idx] if tags_list else ""
        tree.insert("", "end", values=rows[idx], tags=(tag,) if tag else ())
        tree.after(delay, lambda: _insert(idx + 1))
    _insert(0)

# ════════════════════════════════════════════════════════════════════════════
#  📋  КОПИРОВАТЬ СТРОКУ
# ════════════════════════════════════════════════════════════════════════════

def attach_context_menu(tree, toast):
    import tkinter as tk
    ctx = tk.Menu(tree, tearoff=0, bg=CLR_SURFACE2, fg=CLR_TEXT,
                  activebackground=CLR_ACCENT, activeforeground="#ffffff",
                  bd=0, relief="flat", font=("Bahnschrift", 12))

    def copy_cell():
        sel = tree.selection()
        if not sel: return
        val = tree.item(sel[0], "values")
        root.clipboard_clear(); root.clipboard_append(str(val[0]) if val else "")
        toast.show("Скопировано", icon="📋")

    def copy_row():
        sel = tree.selection()
        if not sel: return
        val = tree.item(sel[0], "values")
        root.clipboard_clear(); root.clipboard_append("\t".join(str(v) for v in val))
        toast.show("Строка скопирована", icon="📋")

    def copy_all():
        rows = ["\t".join(str(v) for v in tree.item(k, "values")) for k in tree.get_children()]
        root.clipboard_clear(); root.clipboard_append("\n".join(rows))
        toast.show(f"Скопировано {len(rows)} строк", icon="📋")

    ctx.add_command(label="📋  Копировать первый столбец", command=copy_cell)
    ctx.add_command(label="📄  Копировать всю строку",     command=copy_row)
    ctx.add_separator()
    ctx.add_command(label="📑  Копировать всю таблицу",    command=copy_all)

    def show_menu(event):
        item = tree.identify_row(event.y)
        if item: tree.selection_set(item)
        try: ctx.tk_popup(event.x_root, event.y_root)
        finally: ctx.grab_release()

    tree.bind("<Button-3>", show_menu)

# ════════════════════════════════════════════════════════════════════════════
#  📤  ЭКСПОРТ В EXCEL
# ════════════════════════════════════════════════════════════════════════════

def export_to_excel(tree, sheet_name="Результаты", toast=None):
    rows = [tree.item(k, "values") for k in tree.get_children()]
    if not rows:
        messagebox.showwarning("Экспорт", "Таблица пустая."); return

    columns  = [tree.heading(c)["text"] for c in tree["columns"]]
    save_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx", filetypes=[("Excel файл", "*.xlsx")],
        title="Сохранить результаты",
        initialfile=f"СФМ_{sheet_name}_{datetime.now().strftime('%d.%m.%Y')}.xlsx")
    if not save_path: return

    if HAS_OPENPYXL:
        _export_styled(rows, columns, save_path, sheet_name, tree)
    else:
        pd.DataFrame(rows, columns=columns).to_excel(save_path, index=False, sheet_name=sheet_name)

    if toast: toast.show(f"Экспортировано {len(rows)} строк", icon="📤")
    else: messagebox.showinfo("Экспорт", f"Сохранено: {save_path}")


def _export_styled(rows, columns, path, sheet_name, tree):
    wb = Workbook(); ws = wb.active; ws.title = sheet_name[:31]
    HDR_FILL = PatternFill("solid", fgColor="E2EAF3")
    RED_FILL = PatternFill("solid", fgColor="F0D8D8")
    GRN_FILL = PatternFill("solid", fgColor="D0EEDD")
    YLW_FILL = PatternFill("solid", fgColor="F0EAD0")
    ALT_FILL = PatternFill("solid", fgColor="EEF2F7")
    DEF_FILL = PatternFill("solid", fgColor="F8FAFC")
    HDR_FONT = Font(name="Bahnschrift", bold=True, color="6A8090", size=11)
    DEF_FONT = Font(name="Bahnschrift", color="2A3A46", size=11)
    RED_FONT = Font(name="Bahnschrift", color="A03030", size=11)
    GRN_FONT = Font(name="Bahnschrift", color="2A7A48", size=11)
    YLW_FONT = Font(name="Bahnschrift", color="806020", size=11)
    border   = Border(bottom=Side(style="thin", color="C4D4E4"))

    for ci, col in enumerate(columns, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border
    ws.row_dimensions[1].height = 22

    PRP_FILL = PatternFill("solid", fgColor="E8E0F0")
    PRP_FONT = Font(name="Bahnschrift", color="604090", size=11)

    all_children = list(tree.get_children())
    for ri, rv in enumerate(rows, 2):
        tag    = tree.item(all_children[ri - 2], "tags")
        is_red = "red" in tag or "excluded" in tag or "revoked" in tag
        is_grn = "active" in tag
        is_ylw = "restricted" in tag or "cancelled" in tag
        is_prp = "liquidated" in tag
        for ci, val in enumerate(rv, 1):
            c = ws.cell(row=ri, column=ci, value=str(val))
            c.fill = (RED_FILL if is_red else GRN_FILL if is_grn else
                      PRP_FILL if is_prp else YLW_FILL if is_ylw else
                      (ALT_FILL if ri%2==0 else DEF_FILL))
            c.font = (RED_FONT if is_red else GRN_FONT if is_grn else
                      PRP_FONT if is_prp else YLW_FONT if is_ylw else DEF_FONT)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = border
        ws.row_dimensions[ri].height = 20

    for ci, col in enumerate(columns, 1):
        mw = max([len(str(r[ci-1])) for r in rows] + [len(col)]) + 4
        ws.column_dimensions[get_column_letter(ci)].width = min(mw, 60)
    ws.freeze_panes = "A2"
    wb.save(path)

# ════════════════════════════════════════════════════════════════════════════
#  🔍  РАСШИРЕННЫЙ ПОИСК
# ════════════════════════════════════════════════════════════════════════════

class AdvancedSearch:
    def __init__(self, parent, tree, all_rows_ref):
        self.tree         = tree
        self.all_rows_ref = all_rows_ref

        self.frame = ctk.CTkFrame(parent, fg_color=CLR_SURFACE2,
                                  corner_radius=10, border_width=1, border_color=CLR_BORDER)
        row1 = ctk.CTkFrame(self.frame, fg_color="transparent")
        row1.pack(fill="x", padx=10, pady=(8, 4))

        self.search_var = ctk.StringVar()
        self.search_var.trace_add("write", lambda *_: self.apply())
        ctk.CTkEntry(row1, textvariable=self.search_var, width=260,
                     fg_color=CLR_SURFACE, border_color=CLR_BORDER, text_color=CLR_TEXT,
                     placeholder_text="🔍 Поиск...", corner_radius=8,
                     font=("Bahnschrift", 12)).pack(side="left", padx=(0, 8))

        self.status_var = ctk.StringVar(value="Все статусы")
        self.status_cb  = ctk.CTkComboBox(row1, values=["Все статусы"],
                                          variable=self.status_var, width=200,
                                          fg_color=CLR_SURFACE, border_color=CLR_BORDER,
                                          button_color=CLR_ACCENT, text_color=CLR_TEXT,
                                          dropdown_fg_color=CLR_SURFACE2,
                                          font=("Bahnschrift", 12), corner_radius=8,
                                          command=lambda _: self.apply())
        self.status_cb.pack(side="left", padx=(0, 8))

        ctk.CTkButton(row1, text="✖ Сброс", width=90, height=32,
                      fg_color=CLR_SURFACE, hover_color=CLR_BORDER,
                      text_color=CLR_MUTED, border_width=1, border_color=CLR_BORDER,
                      corner_radius=8, font=("Bahnschrift", 12),
                      command=self.reset).pack(side="left")

        self.count_lbl = ctk.CTkLabel(self.frame, text="",
                                      font=("Bahnschrift", 11), text_color=CLR_MUTED)
        self.count_lbl.pack(anchor="w", padx=12, pady=(0, 8))

    def pack(self, **kwargs):
        self.frame.pack(**kwargs)

    def update_statuses(self):
        statuses = set()
        for r in self.all_rows_ref:
            if len(r["values"]) >= 3:
                statuses.add(str(r["values"][2]))
        self.status_cb.configure(values=["Все статусы"] + sorted(statuses))

    def apply(self):
        query  = normalize(self.search_var.get())
        status = self.status_var.get()
        for item in self.tree.get_children():
            self.tree.delete(item)
        show = 0
        for rd in self.all_rows_ref:
            vals   = rd["values"]
            tag    = rd["tag"]
            fio_ok = (not query) or (query in normalize(str(vals[0])))
            st_ok  = (status == "Все статусы") or (len(vals) >= 3 and str(vals[2]) == status)
            if fio_ok and st_ok:
                self.tree.insert("", "end", values=vals, tags=(tag,) if tag else ())
                show += 1
        total = len(self.all_rows_ref)
        self.count_lbl.configure(
            text=f"Показано: {show} из {total}" if (query or status != "Все статусы") else f"Всего: {total}")

    def reset(self):
        self.search_var.set("")
        self.status_var.set("Все статусы")
        self.apply()

# ════════════════════════════════════════════════════════════════════════════
#  📊  ДАШБОРД
# ════════════════════════════════════════════════════════════════════════════

COLOR_MAP = {
    # ── Террористы ──────────────────────────────────
    "В перечне":        "#f85149",   # красный
    "Нет в перечне":    "#3fb950",   # зелёный
    "Исключен":         "#f85149",   # красный (опасный статус)
    "Добавлен":         "#f85149",   # красный
    "Удален":           "#d29922",   # жёлтый
    # ── МФО ─────────────────────────────────────────
    "Действующий":      "#3fb950",   # зелёный
    "Исключён":         "#f85149",   # красный (исключён из реестра = опасно)
    "Не найден":        "#484f58",   # серый
    # ── Банки (реальные статусы из файла ЦБ) ────────
    "Действующая":      "#3fb950",   # зелёный
    "Отозванная":       "#f85149",   # красный
    "Аннулированная":   "#d29922",   # жёлтый
    "Ликвидация":       "#a371f7",   # фиолетовый
    "nan":              "#484f58",   # серый (пустые строки)
}

def get_status_color(status: str) -> str:
    """Цвет по статусу — сначала точное совпадение, потом по ключевым словам."""
    # Точное совпадение
    if status in COLOR_MAP:
        return COLOR_MAP[status]
    # Поиск по ключевым словам (на случай длинных строк от ЦБ)
    s = status.upper()
    if "ДЕЙСТВУ" in s:               return "#3fb950"  # зелёный
    if "ОТОЗВАН" in s:               return "#f85149"  # красный
    if "АННУЛИРОВАН" in s:           return "#d29922"  # жёлтый
    if "ЛИКВИДАЦ" in s or "ЛИКВИДИР" in s: return "#8070b0"  # фиолетовый
    if "ЗАПРЕЩ" in s or "ОГРАНИЧЕН" in s: return "#b08030"  # жёлтый приглушённый
    if "ИСКЛЮЧ" in s:                return "#c05a5a"  # красный приглушённый
    if "В ПЕРЕЧНЕ" in s:             return "#c05a5a"  # красный приглушённый
    if "НЕТ В ПЕРЕЧНЕ" in s:         return "#4aaa6a"  # зелёный
    if "НЕ НАЙДЕН" in s:             return "#8090a0"  # серый
    return "#6a8090"  # серый по умолчанию

def open_dashboard(all_rows, title="Статистика проверки"):
    if not all_rows:
        messagebox.showinfo("Дашборд", "Нет данных — сначала выполните проверку."); return

    dw = ctk.CTkToplevel(root)
    dw.title(f"📊 Дашборд — {title}")
    dw.geometry("880x600")
    dw.configure(fg_color=CLR_BG)
    dw.lift(); dw.attributes("-topmost", True)
    dw.after(150, lambda: dw.attributes("-topmost", False))

    hdr = ctk.CTkFrame(dw, fg_color=CLR_SURFACE, corner_radius=0, height=52)
    hdr.pack(fill="x"); hdr.pack_propagate(False)
    ctk.CTkLabel(hdr, text=f"📊  {title}",
                 font=("Bahnschrift", 15, "bold"), text_color=CLR_TEXT).pack(side="left", padx=20, pady=12)

    sf = ctk.CTkScrollableFrame(dw, fg_color=CLR_BG)
    sf.pack(fill="both", expand=True, padx=16, pady=12)

    statuses = [str(r["values"][2]) if len(r["values"]) >= 3 else "—" for r in all_rows]
    counts   = Counter(statuses)
    total    = len(all_rows)

    # KPI
    kpi_frame = ctk.CTkFrame(sf, fg_color="transparent")
    kpi_frame.pack(fill="x", pady=(0, 16))
    kpi_data = [("Всего записей", str(total), CLR_ACCENT, "📋")]
    for status, count in counts.most_common(3):
        color = get_status_color(status)
        kpi_data.append((status, str(count), color, "●"))
    for label, value, color, icon in kpi_data:
        card = ctk.CTkFrame(kpi_frame, fg_color=CLR_SURFACE, corner_radius=12,
                            border_width=1, border_color=CLR_BORDER)
        card.pack(side="left", expand=True, fill="both", padx=6)
        ctk.CTkLabel(card, text=icon, font=("Bahnschrift", 22)).pack(pady=(14, 2))
        ctk.CTkLabel(card, text=value, font=("Bahnschrift", 28, "bold"), text_color=color).pack()
        ctk.CTkLabel(card, text=label, font=("Bahnschrift", 10), text_color=CLR_MUTED,
                     wraplength=140).pack(pady=(2, 14))

    import tkinter as tk

    # Bar chart
    bar_card = ctk.CTkFrame(sf, fg_color=CLR_SURFACE, corner_radius=12,
                            border_width=1, border_color=CLR_BORDER)
    bar_card.pack(fill="x", pady=(0, 16))
    ctk.CTkLabel(bar_card, text="Распределение по статусам",
                 font=("Bahnschrift", 13, "bold"), text_color=CLR_TEXT).pack(anchor="w", padx=18, pady=(14, 8))
    bar_canvas = tk.Canvas(bar_card, bg=CLR_SURFACE, bd=0, highlightthickness=0, height=200)
    bar_canvas.pack(fill="x", padx=18, pady=(0, 16))

    def draw_bars(event=None):
        bar_canvas.delete("all")
        items = [(k, v) for k, v in counts.items() if v > 0]
        if not items: return
        w = bar_canvas.winfo_width() or 800
        max_v, bar_h, gap, label_w = max(v for _,v in items), 28, 14, 220
        for i, (status, count) in enumerate(items):
            y     = 10 + i * (bar_h + gap)
            bw    = int((w - label_w - 80) * count / max_v)
            color = get_status_color(status)
            bar_canvas.create_text(label_w - 8, y + bar_h//2, text=status,
                                   anchor="e", fill=CLR_MUTED, font=("Bahnschrift", 11))
            bar_canvas.create_rectangle(label_w, y, label_w + (w - label_w - 80), y + bar_h,
                                        fill=CLR_SURFACE2, outline="")
            if bw > 0:
                bar_canvas.create_rectangle(label_w, y, label_w + bw, y + bar_h, fill=color, outline="")
            bar_canvas.create_text(label_w + bw + 8, y + bar_h//2, text=str(count),
                                   anchor="w", fill=CLR_TEXT, font=("Bahnschrift", 11, "bold"))
        bar_canvas.configure(height=10 + len(items) * (bar_h + gap))

    bar_canvas.bind("<Configure>", draw_bars)
    dw.after(100, draw_bars)

    # Pie chart
    pie_card = ctk.CTkFrame(sf, fg_color=CLR_SURFACE, corner_radius=12,
                            border_width=1, border_color=CLR_BORDER)
    pie_card.pack(fill="x", pady=(0, 16))
    ctk.CTkLabel(pie_card, text="Доля по статусам",
                 font=("Bahnschrift", 13, "bold"), text_color=CLR_TEXT).pack(anchor="w", padx=18, pady=(14, 8))
    pie_canvas = tk.Canvas(pie_card, bg=CLR_SURFACE, bd=0, highlightthickness=0, height=230)
    pie_canvas.pack(fill="x", padx=18, pady=(0, 16))

    def draw_pie(event=None):
        pie_canvas.delete("all")
        items = [(k, v) for k, v in counts.items() if v > 0]
        if not items: return
        cx, cy, r, start = 130, 110, 90, 0
        for status, count in items:
            extent = 360 * count / total
            color  = get_status_color(status)
            pie_canvas.create_arc(cx-r, cy-r, cx+r, cy+r, start=start, extent=extent,
                                  fill=color, outline=CLR_BG, width=2)
            start += extent
        lx, ly = cx + r + 30, cy - len(items) * 14
        for status, count in items:
            color = get_status_color(status)
            pie_canvas.create_rectangle(lx, ly, lx+14, ly+14, fill=color, outline="")
            pie_canvas.create_text(lx+20, ly+7, text=f"{status}  {count}  ({count/total*100:.1f}%)",
                                   anchor="w", fill=CLR_TEXT, font=("Bahnschrift", 11))
            ly += 24

    pie_canvas.bind("<Configure>", draw_pie)
    dw.after(150, draw_pie)

    # Время
    info_card = ctk.CTkFrame(sf, fg_color=CLR_SURFACE, corner_radius=12,
                              border_width=1, border_color=CLR_BORDER)
    info_card.pack(fill="x", pady=(0, 8))
    ctk.CTkLabel(info_card,
                 text=f"🕐  Отчёт сформирован: {datetime.now().strftime('%d.%m.%Y  %H:%M:%S')}",
                 font=("Bahnschrift", 12), text_color=CLR_MUTED).pack(anchor="w", padx=18, pady=12)

# ════════════════════════════════════════════════════════════════════════════
#  СОРТИРОВКА
# ════════════════════════════════════════════════════════════════════════════

def sort_column(tv, col, reverse):
    data = [(tv.set(k, col), k) for k in tv.get_children("")]
    data.sort(reverse=reverse)
    for i, (_, k) in enumerate(data): tv.move(k, "", i)
    tv.heading(col, command=lambda: sort_column(tv, col, not reverse))

# ════════════════════════════════════════════════════════════════════════════
#  ИСТОРИЯ ИЗМЕНЕНИЙ
# ════════════════════════════════════════════════════════════════════════════

def open_history_window():
    w = ctk.CTkToplevel(root)
    w.title("История изменений")
    w.geometry("760x580")
    w.configure(fg_color=CLR_BG)
    ctk.CTkLabel(w, text="📋  История изменений",
                 font=("Bahnschrift", 16, "bold"), text_color=CLR_TEXT).pack(padx=20, pady=(16,8), anchor="w")
    ctk.CTkFrame(w, height=1, fg_color=CLR_BORDER).pack(fill="x", padx=20, pady=(0,8))
    text = ctk.CTkTextbox(w, wrap="word", font=("Bahnschrift", 12),
                          fg_color=CLR_SURFACE, text_color=CLR_TEXT,
                          border_color=CLR_BORDER, border_width=1, corner_radius=10)
    text.pack(expand=True, fill="both", padx=20, pady=(0,16))
    text.insert("0.0", """
Минимальные требования:
Python 3.11–3.14, pip
Библиотеки: customtkinter, pandas, openpyxl, requests

СФМ — История версий

Версия 1.0 – Проверяет имена и фамилии в XML-файлах, заложена основная логика.
Версия 1.1 – Появился удобный интерфейс с таблицей и подсветкой ошибок.
Версия 1.2 – Добавлены блоки «Последние исключенные» и «Актуальный перечень».
Версия 1.3 – Сравнение списков: кто добавлен, кто удалён.
Версия 1.4 – Добавлена дата рождения, автоматически подстраивается ширина колонок.
Версия 1.5 – История изменений, статистика
Версия 1.6 – Добавлена проверка МФО через реестр ЦБ РФ, добавлены анимации, можно копировать данные, экспортировать в Excel, расширенный поиск и дашборд.
Версия 1.7 – Добавлена проверка банков через реестр ЦБ РФ и обновлённый дизайн.
""")
    text.configure(state="disabled")

# ════════════════════════════════════════════════════════════════════════════
#  ГЛАВНОЕ МЕНЮ
# ════════════════════════════════════════════════════════════════════════════

def main_menu():
    root.geometry("560x560")
    clear_frame()
    frame = ctk.CTkFrame(root, fg_color=CLR_BG)
    frame.pack(expand=True, fill="both")
    global current_frame
    current_frame = frame

    title_frame = ctk.CTkFrame(frame, fg_color=CLR_SURFACE, corner_radius=14,
                               border_width=1, border_color=CLR_BORDER)
    title_frame.pack(fill="x", padx=28, pady=(24, 18))

    ctk.CTkLabel(title_frame, text="СФМ",
                 font=("Bahnschrift", 32, "bold"), text_color=CLR_ACCENT).pack(pady=(20, 2))
    ctk.CTkLabel(title_frame, text="Служба финансового мониторинга  v1.7",
                 font=("Bahnschrift", 11), text_color=CLR_MUTED).pack(pady=(0, 6))
    ctk.CTkFrame(title_frame, height=1, fg_color=CLR_BORDER).pack(fill="x")
    ctk.CTkLabel(title_frame,
                 text="Где бы ни скрывалась финансовая тайна – она будет раскрыта",
                 font=("Bahnschrift", 10), text_color=CLR_MUTED,
                 wraplength=440).pack(pady=(8, 16))

    btn_style = {"width": 380, "height": 46, "corner_radius": 12,
                 "fg_color": CLR_SURFACE2, "hover_color": CLR_HIGHLIGHT,
                 "text_color": CLR_TEXT, "font": ("Bahnschrift", 14),
                 "anchor": "w", "border_width": 1, "border_color": CLR_BORDER}

    nav = ctk.CTkFrame(frame, fg_color=CLR_BG)
    nav.pack(pady=4)

    ctk.CTkButton(nav, text="🔥   Террористы",
                  command=open_terrorists_menu, **btn_style).pack(pady=5)
    ctk.CTkButton(nav, text="🏦   Проверка МФО (МКК / МФК)",
                  command=open_mfo_check_window, **btn_style).pack(pady=5)
    ctk.CTkButton(nav, text="🏦   Проверка Банков",
                  command=open_banks_check_window, **btn_style).pack(pady=5)
    ctk.CTkButton(nav, text="📜   История изменений",
                  command=open_history_window, **btn_style).pack(pady=5)
    ctk.CTkButton(nav, text="🚀   Будущее СФМ",
                  command=lambda: messagebox.showinfo("Информация", "В стадии разработки"),
                  **btn_style).pack(pady=5)
    ctk.CTkButton(nav, text="Выйти", command=root.destroy,
                  fg_color=CLR_SURFACE2, hover_color="#f0ddd8",
                  text_color=CLR_DANGER, border_color="#d4a0a0", border_width=1,
                  width=380, height=46, corner_radius=12,
                  font=("Bahnschrift", 14), anchor="w").pack(pady=(12, 4))

# ════════════════════════════════════════════════════════════════════════════
#  🏛️  ОКНО БАНКИ
# ════════════════════════════════════════════════════════════════════════════

def open_banks_check_window():
    root.geometry("1400x820")
    clear_frame()
    frame = ctk.CTkFrame(root, fg_color=CLR_BG)
    frame.pack(fill="both", expand=True)
    global current_frame
    current_frame = frame

    # Шапка с датой запроса
    hdr = ctk.CTkFrame(frame, fg_color=CLR_SURFACE, corner_radius=0, height=52)
    hdr.pack(fill="x"); hdr.pack_propagate(False)
    today_str = datetime.now().strftime("%d.%m.%Y")
    ctk.CTkLabel(hdr, text=f"🏛️  Проверка Банков по реестру ЦБ РФ  —  {today_str}",
                 font=("Bahnschrift", 15, "bold"), text_color=CLR_TEXT).pack(side="left", padx=20, pady=12)

    # Колонки: ОГРН (из локального файла) | Наименование (ЦБ) | Статус лицензии (ЦБ)
    columns = ("ОГРН", "Наименование", "Статус лицензии")
    tree = ttk.Treeview(frame, columns=columns, show="headings", height=20)
    tree.heading("ОГРН",              text="ОГРН",              command=lambda: sort_column(tree, "ОГРН",              False))
    tree.heading("Наименование",      text="Наименование",      command=lambda: sort_column(tree, "Наименование",      False))
    tree.heading("Статус лицензии",   text="Статус лицензии",   command=lambda: sort_column(tree, "Статус лицензии",   False))
    tree.column("ОГРН",             anchor="w", width=180)
    tree.column("Наименование",     anchor="w", width=420)
    tree.column("Статус лицензии",  anchor="w", width=320)

    # Цвета тегов
    tree.tag_configure("active",     background="#d0eedd", foreground="#2a7a48")   # лицензия действует
    tree.tag_configure("revoked",    background="#f0d8d8", foreground="#a03030")   # отозвана
    tree.tag_configure("cancelled",  background="#f0ead0", foreground="#806020")   # аннулирована
    tree.tag_configure("liquidated", background="#e8e0f0", foreground="#604090")   # ликвидация
    tree.tag_configure("restricted", background="#f0e8d0", foreground="#805010")   # ограничения / запрет
    tree.tag_configure("notfound",   background=CLR_SURFACE, foreground=CLR_MUTED)

    scroll_y = ctk.CTkScrollbar(frame, orientation="vertical", command=tree.yview,
                                button_color=CLR_ACCENT, button_hover_color=CLR_ACCENT2)
    tree.configure(yscrollcommand=scroll_y.set)

    all_rows   = []
    adv_search = AdvancedSearch(frame, tree, all_rows)
    adv_search.pack(fill="x", padx=12, pady=4)

    overlay = LoadingOverlay(frame)
    toast   = ToastNotification(frame)
    attach_context_menu(tree, toast)

    # Кнопки
    bf = ctk.CTkFrame(frame, fg_color=CLR_BG); bf.pack(pady=8)
    _b = {"corner_radius": 12, "font": ("Bahnschrift", 13),
          "fg_color": CLR_SURFACE2, "hover_color": CLR_HIGHLIGHT,
          "text_color": CLR_TEXT, "border_width": 1, "border_color": CLR_BORDER}

    ctk.CTkButton(bf, text="🔄  Проверить", width=190,
                  command=lambda: check_banks(tree, overlay, toast, all_rows, adv_search),
                  fg_color=CLR_ACCENT, hover_color=CLR_ACCENT2,
                  corner_radius=8, font=("Bahnschrift", 13)).pack(side="left", padx=6)
    ctk.CTkButton(bf, text="📤  Экспорт Excel", width=180,
                  command=lambda: export_to_excel(tree, "Банки", toast), **_b).pack(side="left", padx=6)
    ctk.CTkButton(bf, text="📊  Дашборд", width=150,
                  command=lambda: open_dashboard(all_rows, "Банки — статистика"), **_b).pack(side="left", padx=6)
    ctk.CTkButton(bf, text="⬅  Назад", command=main_menu, width=140, **_b).pack(side="left", padx=6)
    ctk.CTkButton(bf, text="Выйти", command=root.destroy,
                  fg_color=CLR_SURFACE2, hover_color="#f0ddd8",
                  text_color=CLR_DANGER, border_color="#d4a0a0", border_width=1,
                  corner_radius=8, width=120, font=("Bahnschrift", 13)).pack(side="left", padx=6)

    tree.pack(side="left", fill="both", expand=True, padx=(12, 0), pady=8)
    scroll_y.pack(side="right", fill="y", pady=8, padx=(0, 4))

# ════════════════════════════════════════════════════════════════════════════
#  🏛️  ЛОГИКА — check_banks
# ════════════════════════════════════════════════════════════════════════════

def _get_bank_tag(status: str) -> str:
    """Определяем тег по тексту статуса лицензии."""
    s = status.upper()
    if "ДЕЙСТВУЕТ" in s or "ДЕЙСТВУЮЩ" in s:
        return "active"
    if "ОТОЗВАНА" in s or "ОТОЗВАН" in s:
        return "revoked"
    if "АННУЛИРОВАН" in s:
        return "cancelled"
    if "ЛИКВИДАЦ" in s or "ЛИКВИДИР" in s:
        return "liquidated"
    if "ЗАПРЕЩ" in s or "ОГРАНИЧЕН" in s or "ПРИНУДИТЕЛЬН" in s:
        return "restricted"
    return "notfound"


def check_banks(tree, overlay, toast, all_rows, adv_search):
    tree.delete(*tree.get_children())
    all_rows.clear()
    overlay.show("Подключение к ЦБ РФ...", "Скачивание реестра банков")

    def worker():
        # ── 1. Скачиваем реестр ЦБ ─────────────────────────────────────────
        url = get_banks_cbr_url()
        overlay.set_progress(0.1, "Скачивание реестра...",
                             f"Дата: {datetime.now().strftime('%d.%m.%Y')}")
        try:
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                              "AppleWebKit/537.36 (KHTML, like Gecko) "
                              "Chrome/120.0.0.0 Safari/537.36",
                "Referer": "https://www.cbr.ru/banking_sector/credit/FullCoList/",
            }
            resp = requests.get(url, headers=headers, timeout=40)
            resp.raise_for_status()
            cbr_path = os.path.join(os.getenv("TEMP"), "cbr_banks.xlsx")
            with open(cbr_path, "wb") as f:
                f.write(resp.content)
        except Exception as e:
            root.after(0, overlay.hide)
            root.after(0, lambda: messagebox.showerror(
                "Ошибка", f"Не удалось скачать реестр банков с ЦБ:\n{e}\n\nURL: {url}"))
            return

        # ── 2. Читаем реестр ЦБ ────────────────────────────────────────────
        overlay.set_progress(0.3, "Чтение реестра ЦБ...", "")
        try:
            # Файл ЦБ содержит строки-заголовки вверху — читаем без header
            # Столбец D (индекс 3) = ОГРН, E (4) = Наименование, H (7) = Статус лицензии
            cbr_df = pd.read_excel(cbr_path, header=None, dtype=str)
        except Exception as e:
            root.after(0, overlay.hide)
            root.after(0, lambda: messagebox.showerror("Ошибка", f"Ошибка чтения файла ЦБ:\n{e}"))
            return

        # Ищем строку-заголовок (где в столбце D есть "огрн" или "рег")
        header_row = 0
        for i, row in cbr_df.iterrows():
            cell = str(row.iloc[3]).lower()
            if "огрн" in cell or "рег" in cell:
                header_row = i
                break

        # Данные начинаются со следующей строки после заголовка
        data_df = cbr_df.iloc[header_row + 1:].reset_index(drop=True)

        # Строим словарь ОГРН → (Наименование, Статус)
        cbr_dict = {}
        for _, row in data_df.iterrows():
            ogrn   = clean_ogrn(row.iloc[3])   # D
            name   = str(row.iloc[4]).strip()   # E
            status = str(row.iloc[7]).strip()   # H
            if ogrn and ogrn not in ("nan", "None", ""):
                cbr_dict[ogrn] = (name, status)

        # ── 3. Читаем локальный файл (лист "Банки", столбец A) ────────────
        overlay.set_progress(0.6, "Чтение локального файла...",
                             f"Лист 'Банки', столбец A")
        try:
            local_df = pd.read_excel(MFO_LOCAL_PATH, sheet_name="Банки",
                                     header=None, dtype=str)
        except Exception as e:
            root.after(0, overlay.hide)
            root.after(0, lambda: messagebox.showerror(
                "Ошибка",
                f"Не удалось открыть лист 'Банки' из файла:\n{MFO_LOCAL_PATH}\n\n{e}"))
            return

        # ── 4. Сверка ──────────────────────────────────────────────────────
        overlay.set_progress(0.8, "Сверка данных...", f"Реестр ЦБ: {len(cbr_dict)} банков")
        rows, tags = [], []

        for _, row in local_df.iterrows():
            ogrn_local = clean_ogrn(row.iloc[0])   # столбец A
            if not ogrn_local or ogrn_local in ("nan", "None"):
                continue

            if ogrn_local in cbr_dict:
                name, status = cbr_dict[ogrn_local]
                tag = _get_bank_tag(status)
                rows.append((ogrn_local, name, status))
                tags.append(tag)
            else:
                rows.append((ogrn_local, "— не найден в реестре ЦБ —", "Не найден"))
                tags.append("notfound")

        overlay.set_progress(1.0, "Готово!", "")

        def finish():
            overlay.hide()
            for r, t in zip(rows, tags):
                all_rows.append({"values": r, "tag": t})
            adv_search.update_statuses()
            animate_rows(tree, rows, tags, delay=20)
            root.after(400, lambda: auto_resize(tree))
            toast.show(f"Проверено {len(rows)} банков", icon="🏛️")

        root.after(400, finish)

    threading.Thread(target=worker, daemon=True).start()

# ════════════════════════════════════════════════════════════════════════════
#  ТЕРРОРИСТЫ
# ════════════════════════════════════════════════════════════════════════════

def open_terrorists_menu():
    root.geometry("1350x800")
    clear_frame()
    frame = ctk.CTkFrame(root, fg_color=CLR_BG)
    frame.pack(fill="both", expand=True)
    global current_frame
    current_frame = frame

    hdr = ctk.CTkFrame(frame, fg_color=CLR_SURFACE, corner_radius=0, height=52)
    hdr.pack(fill="x"); hdr.pack_propagate(False)
    ctk.CTkLabel(hdr, text="🔥  Проверка по перечню террористов",
                 font=("Bahnschrift", 15, "bold"), text_color=CLR_TEXT).pack(side="left", padx=20, pady=12)

    columns = ("ФИО", "Дата рождения", "Статус", "Последняя дата", "Изменение")
    tree = ttk.Treeview(frame, columns=columns, show="headings", height=20)
    for col in columns:
        tree.heading(col, text=col, command=lambda c=col: sort_column(tree, c, False))
        tree.column(col, anchor="w", width=200)
    tree.tag_configure("red",   background="#f0d8d8", foreground="#a03030")
    tree.tag_configure("found", background=CLR_HIGHLIGHT, foreground=CLR_ACCENT)

    scroll_y = ctk.CTkScrollbar(frame, orientation="vertical", command=tree.yview,
                                button_color=CLR_ACCENT, button_hover_color=CLR_ACCENT2)
    tree.configure(yscrollcommand=scroll_y.set)
    all_rows = []

    sf = ctk.CTkFrame(frame, fg_color=CLR_SURFACE2, corner_radius=10, height=42)
    sf.pack(fill="x", padx=12, pady=(8,4)); sf.pack_propagate(False)
    label_in    = ctk.CTkLabel(sf, text="В перечне: 0",    font=("Bahnschrift", 12), text_color=CLR_DANGER)
    label_not   = ctk.CTkLabel(sf, text="Нет в перечне: 0", font=("Bahnschrift", 12), text_color=CLR_SUCCESS)
    label_excl  = ctk.CTkLabel(sf, text="Исключён: 0",     font=("Bahnschrift", 12), text_color=CLR_MUTED)
    label_in.pack(side="left", padx=18); label_not.pack(side="left", padx=18); label_excl.pack(side="left", padx=18)

    adv_search = AdvancedSearch(frame, tree, all_rows)
    adv_search.pack(fill="x", padx=12, pady=4)

    overlay = LoadingOverlay(frame)
    toast   = ToastNotification(frame)
    attach_context_menu(tree, toast)

    bf = ctk.CTkFrame(frame, fg_color=CLR_BG); bf.pack(pady=6)
    _b = {"corner_radius": 12, "width": 185, "font": ("Bahnschrift", 13),
          "fg_color": CLR_SURFACE2, "hover_color": CLR_HIGHLIGHT,
          "text_color": CLR_TEXT, "border_width": 1, "border_color": CLR_BORDER}

    ctk.CTkButton(bf, text="📂  Проверить изменения",
                  command=lambda: check_xml(tree, label_in, label_not, label_excl,
                                            overlay, toast, all_rows, adv_search),
                  fg_color=CLR_ACCENT, hover_color=CLR_ACCENT2, border_width=0,
                  corner_radius=8, width=185, font=("Bahnschrift", 13)).pack(side="left", padx=5)
    ctk.CTkButton(bf, text="🔎  Сверка перечней",
                  command=lambda: compare_lists(tree, overlay, toast, all_rows, adv_search),
                  **_b).pack(side="left", padx=5)
    ctk.CTkButton(bf, text="📤  Экспорт Excel",
                  command=lambda: export_to_excel(tree, "Террористы", toast), **_b).pack(side="left", padx=5)
    ctk.CTkButton(bf, text="📊  Дашборд",
                  command=lambda: open_dashboard(all_rows, "Террористы — статистика"), **_b).pack(side="left", padx=5)
    ctk.CTkButton(bf, text="⬅  Назад", command=main_menu, **_b).pack(side="left", padx=5)
    ctk.CTkButton(bf, text="Выйти", command=root.destroy,
                  fg_color=CLR_SURFACE2, hover_color="#f0ddd8",
                  text_color=CLR_DANGER, border_color="#d4a0a0", border_width=1,
                  corner_radius=8, width=110, font=("Bahnschrift", 13)).pack(side="left", padx=5)

    tree.pack(side="left", fill="both", expand=True, padx=(12, 0), pady=8)
    scroll_y.pack(side="right", fill="y", pady=8, padx=(0, 4))

# ════════════════════════════════════════════════════════════════════════════
#  МФО
# ════════════════════════════════════════════════════════════════════════════

def open_mfo_check_window():
    root.geometry("1350x800")
    clear_frame()
    frame = ctk.CTkFrame(root, fg_color=CLR_BG)
    frame.pack(fill="both", expand=True)
    global current_frame
    current_frame = frame

    hdr = ctk.CTkFrame(frame, fg_color=CLR_SURFACE, corner_radius=0, height=52)
    hdr.pack(fill="x"); hdr.pack_propagate(False)
    ctk.CTkLabel(hdr, text="🏦  Проверка МФО по реестру ЦБ РФ",
                 font=("Bahnschrift", 15, "bold"), text_color=CLR_TEXT).pack(side="left", padx=20, pady=12)

    columns = ("Наименование", "ИНН", "Статус")
    tree = ttk.Treeview(frame, columns=columns, show="headings", height=20)
    for col in columns:
        tree.heading(col, text=col, command=lambda c=col: sort_column(tree, c, False))
        tree.column(col, anchor="w", width=280)
    tree.tag_configure("excluded", background="#f0d8d8", foreground="#a03030")
    tree.tag_configure("active",   background="#d0eedd", foreground="#2a7a48")

    scroll_y = ctk.CTkScrollbar(frame, orientation="vertical", command=tree.yview,
                                button_color=CLR_ACCENT, button_hover_color=CLR_ACCENT2)
    tree.configure(yscrollcommand=scroll_y.set)
    all_rows = []
    adv_search = AdvancedSearch(frame, tree, all_rows)
    adv_search.pack(fill="x", padx=12, pady=4)

    overlay = LoadingOverlay(frame)
    toast   = ToastNotification(frame)
    attach_context_menu(tree, toast)

    bf = ctk.CTkFrame(frame, fg_color=CLR_BG); bf.pack(pady=10)
    _b = {"corner_radius": 12, "font": ("Bahnschrift", 13),
          "fg_color": CLR_SURFACE2, "hover_color": CLR_HIGHLIGHT,
          "text_color": CLR_TEXT, "border_width": 1, "border_color": CLR_BORDER}

    ctk.CTkButton(bf, text="🔄  Проверить", width=180,
                  command=lambda: check_mfo(tree, overlay, toast, all_rows, adv_search),
                  fg_color=CLR_ACCENT, hover_color=CLR_ACCENT2,
                  corner_radius=8, font=("Bahnschrift", 13)).pack(side="left", padx=8)
    ctk.CTkButton(bf, text="📤  Экспорт Excel", width=170,
                  command=lambda: export_to_excel(tree, "МФО", toast), **_b).pack(side="left", padx=8)
    ctk.CTkButton(bf, text="📊  Дашборд", width=150,
                  command=lambda: open_dashboard(all_rows, "МФО — статистика"), **_b).pack(side="left", padx=8)
    ctk.CTkButton(bf, text="⬅  Назад", command=main_menu, width=140, **_b).pack(side="left", padx=8)

    tree.pack(side="left", fill="both", expand=True, padx=(12, 0), pady=8)
    scroll_y.pack(side="right", fill="y", pady=8, padx=(0, 4))

# ════════════════════════════════════════════════════════════════════════════
#  ЛОГИКА — check_mfo
# ════════════════════════════════════════════════════════════════════════════

def check_mfo(tree, overlay, toast, all_rows, adv_search):
    tree.delete(*tree.get_children()); all_rows.clear()
    overlay.show("Подключение к ЦБ РФ...", "Скачивание реестра МФО")

    def worker():
        def clean(v):
            if pd.isna(v): return ""
            return str(v).strip().replace(".0", "").replace(" ", "")

        try:
            overlay.set_progress(0.1, "Скачивание реестра...", "cbr.ru → list_MFO.xlsx")
            r = requests.get(MFO_CBR_URL, timeout=30)
            cbr_path = os.path.join(os.getenv("TEMP"), "list_MFO.xlsx")
            with open(cbr_path, "wb") as f: f.write(r.content)
        except Exception as e:
            root.after(0, overlay.hide)
            root.after(0, lambda: messagebox.showerror("Ошибка", str(e))); return

        overlay.set_progress(0.3, "Чтение локального файла...", "")
        try:
            local_df = pd.read_excel(MFO_LOCAL_PATH, dtype=str)
        except Exception as e:
            root.after(0, overlay.hide)
            root.after(0, lambda: messagebox.showerror("Ошибка", str(e))); return

        overlay.set_progress(0.5, "Обработка реестра ЦБ...", "Действующие")
        cbr_file = pd.ExcelFile(cbr_path)
        active_dict = {}
        for sheet in ["Действующие", "Действующие МФК", "Действующие МКК"]:
            df = pd.read_excel(cbr_file, sheet_name=sheet, dtype=str)
            for row in df.values:
                inn = clean(row[5])
                if inn: active_dict[inn] = str(row[7]).strip()

        overlay.set_progress(0.7, "Обработка реестра ЦБ...", "Исключённые")
        excl_df   = pd.read_excel(cbr_file, sheet_name="Исключенные", dtype=str)
        excl_dict = {}
        for row in excl_df.values:
            inn = clean(row[6])
            if inn: excl_dict[inn] = str(row[8]).strip()

        overlay.set_progress(0.85, "Сверка...", "")
        rows, tags = [], []
        for _, row in local_df.iterrows():
            inn = clean(row.iloc[0])
            if inn in excl_dict:
                rows.append((excl_dict[inn], inn, "Исключён")); tags.append("excluded")
            elif inn in active_dict:
                rows.append((active_dict[inn], inn, "Действующий")); tags.append("active")
            else:
                rows.append(("", inn, "Не найден")); tags.append("")

        overlay.set_progress(1.0, "Готово!", "")

        def finish():
            overlay.hide()
            for r, t in zip(rows, tags): all_rows.append({"values": r, "tag": t})
            adv_search.update_statuses()
            animate_rows(tree, rows, tags)
            root.after(300, lambda: auto_resize(tree))
            toast.show(f"Проверено {len(rows)} МФО", icon="🏦")

        root.after(400, finish)

    threading.Thread(target=worker, daemon=True).start()

# ════════════════════════════════════════════════════════════════════════════
#  ЛОГИКА — check_xml / compare_lists / check_loans
# ════════════════════════════════════════════════════════════════════════════

def check_xml(tree, label_in, label_not, label_excl, overlay, toast, all_rows, adv_search):
    xml_path = filedialog.askopenfilename(title="Загрузите XML файл", filetypes=[("XML files", "*.xml")])
    if not xml_path: return
    overlay.show("Обработка XML...", os.path.basename(xml_path))

    def worker():
        try:
            fn = os.path.basename(xml_path).replace(".xml", "")
            d, m, y = fn.split(".")
            xml_date = datetime(int(y), int(m), int(d)).date()
        except:
            root.after(0, overlay.hide)
            root.after(0, lambda: messagebox.showerror("Ошибка", "Имя файла: DD.MM.YYYY.xml")); return

        overlay.set_progress(0.2, "Чтение Excel...", "")
        try:
            raw = pd.read_excel(BASE_EXCEL_PATH)
            df  = raw.iloc[:, [2, 3]].dropna(subset=[raw.columns[2]])
            df.columns = ["ФИО", "ДатаРождения"]
        except Exception as e:
            root.after(0, overlay.hide)
            root.after(0, lambda: messagebox.showerror("Ошибка", str(e))); return

        overlay.set_progress(0.4, "Маленькая отсылка...", "Бывало, что хотел вклад сделать, а потом появлялись нужды,\nприходилось закрывать. \nБывало, пытался разобраться, и экспериментировал, бывало ошибался. \nБывало полнил не стой карты, проблемы начались, \nнужно было отменять...\nСистема то у вас не очень простая.... Посмотрите мою историю, у меня были у вас вклады, которые весь срок находились у вас....\nПросто сейчас время такое непредсказуемое, вроде хочешь\nсделать хоть небольшой вклад, но что то идет не так.\n— Понамарев Юрий")
        tree_xml = ET.parse(xml_path).getroot()
        excluded, actual = set(), {}
        eb = tree_xml.find("ПоследниеИсключенные")
        if eb:
            for p in eb.iter("ФИО"):
                if p.text: excluded.add(normalize(p.text))
        ab = tree_xml.find("АктуальныйПеречень")
        if ab:
            for subj in ab.findall("Субъект"):
                fl = subj.find("ФЛ")
                if fl is None: continue
                fio = fl.findtext("ФИО")
                if not fio: continue
                dob_xml = format_date(fl.findtext("ДатаРождения") or "")
                hist = subj.find("История"); dates = []
                if hist:
                    for d in hist.findall("ДатаВключения") + hist.findall("ДатаИзменения"):
                        if d.text:
                            p = parse_xml_date(d.text)
                            if p: dates.append(p)
                if dates:
                    key = (normalize(fio), dob_xml)
                    if key not in actual or actual[key] < max(dates):
                        actual[key] = max(dates)

        overlay.set_progress(0.7, "Сравнение...", f"{len(df)} записей")
        rows, tags = [], []
        cnt = {"В перечне": 0, "Нет в перечне": 0, "Исключен": 0}
        for _, row in df.iterrows():
            fio = row["ФИО"]; birth = format_date(row["ДатаРождения"]); norm = normalize(fio)
            if norm in excluded:
                rows.append((fio, birth, "Исключен", xml_date.strftime("%Y-%m-%d"), "ДА"))
                tags.append("red"); cnt["Исключен"] += 1
            else:
                key = (norm, birth)
                if key in actual:
                    ld = actual[key]; ct = ld == xml_date
                    rows.append((fio, birth, "В перечне", ld.strftime("%Y-%m-%d"), "ДА" if ct else "НЕТ"))
                    tags.append("red" if ct else ""); cnt["В перечне"] += 1
                else:
                    rows.append((fio, birth, "Нет в перечне", "", "")); tags.append(""); cnt["Нет в перечне"] += 1

        overlay.set_progress(1.0, "Готово!", "")

        def finish():
            tree.delete(*tree.get_children()); all_rows.clear(); overlay.hide()
            for r, t in zip(rows, tags): all_rows.append({"values": r, "tag": t})
            adv_search.update_statuses()
            animate_rows(tree, rows, tags, delay=12)
            root.after(200, lambda: auto_resize(tree))
            label_in.configure(text=f"В перечне: {cnt['В перечне']}")
            label_not.configure(text=f"Нет в перечне: {cnt['Нет в перечне']}")
            label_excl.configure(text=f"Исключен: {cnt['Исключен']}")
            toast.show(f"Проверено {len(rows)} записей", icon="📂")

        root.after(400, finish)

    threading.Thread(target=worker, daemon=True).start()


def compare_lists(tree, overlay, toast, all_rows, adv_search):
    messagebox.showinfo("Внимание", "Загрузите новый список проверки")
    new_path = filedialog.askopenfilename(title="Новый список", filetypes=[("Excel", "*.xlsx")])
    if not new_path: return
    messagebox.showinfo("Внимание", "Загрузите старый список проверки")
    old_path = filedialog.askopenfilename(title="Старый список", filetypes=[("Excel", "*.xlsx")])
    if not old_path: return
    overlay.show("Сравнение перечней...", "")

    def worker():
        overlay.set_progress(0.3, "Чтение файлов...", "")
        new_df   = pd.read_excel(new_path).iloc[:, [1, 2]].dropna(subset=[pd.read_excel(new_path).columns[1]])
        old_df   = pd.read_excel(old_path).iloc[:, [1, 2]].dropna(subset=[pd.read_excel(old_path).columns[1]])
        new_dict = {normalize(r[0]): r[1] for r in new_df.values}
        old_dict = {normalize(r[0]): r[1] for r in old_df.values}
        overlay.set_progress(0.7, "Поиск изменений...", "")
        rows, tags = [], []
        for n, dr in new_dict.items():
            if n not in old_dict: rows.append((n, dr, "Добавлен", "", "ДА")); tags.append("red")
        for n, dr in old_dict.items():
            if n not in new_dict: rows.append((n, dr, "Удален", "", "ДА")); tags.append("red")
        overlay.set_progress(1.0, "Готово!", "")

        def finish():
            tree.delete(*tree.get_children()); all_rows.clear(); overlay.hide()
            for r, t in zip(rows, tags): all_rows.append({"values": r, "tag": t})
            adv_search.update_statuses()
            animate_rows(tree, rows, tags)
            root.after(200, lambda: auto_resize(tree))
            toast.show(f"Изменений: {len(rows)}", icon="📊")
            ans = messagebox.askyesno("Проверка кредитов", "Проверить наличие выданных кредитов?")
            if ans: check_loans(new_path)

        root.after(400, finish)

    threading.Thread(target=worker, daemon=True).start()


def check_loans(new_path):
    messagebox.showinfo("Загрузка", "Загрузите файл 'Отчет по финансовым сделкам'")
    report_path = filedialog.askopenfilename(title="Отчет по финансовым сделкам", filetypes=[("Excel", "*.xlsx")])
    if not report_path: return
    try:
        new_df = pd.read_excel(new_path, dtype=str).iloc[1:, [0, 1, 2]]
        new_dict = {}
        for _, row in new_df.iterrows():
            id_mpl = str(row.iloc[0]).strip()
            if id_mpl: new_dict[id_mpl] = (str(row.iloc[1]).strip(), row.iloc[2])
        report_df = pd.read_excel(report_path).iloc[2:, :]
        results = []
        for _, row in report_df.iterrows():
            id_mpl = str(row.iloc[0]).strip(); raw = row.iloc[7]
            if id_mpl and pd.notna(raw) and id_mpl in new_dict:
                fio, birth = new_dict[id_mpl]
                results.append((id_mpl, fio, format_date_ru(birth), format_date_ru(raw)))
        if not results:
            messagebox.showinfo("Результат", "Совпадений не найдено"); return

        w = ctk.CTkToplevel(root)
        w.title("Выданные кредиты"); w.geometry("1050x620"); w.configure(fg_color=CLR_BG)
        w.lift(); w.attributes("-topmost", True); w.after(100, lambda: w.attributes("-topmost", False))

        hdr = ctk.CTkFrame(w, fg_color=CLR_SURFACE, corner_radius=0, height=52)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        ctk.CTkLabel(hdr, text="💳  Выданные кредиты — совпадения",
                     font=("Bahnschrift", 15, "bold"), text_color=CLR_TEXT).pack(side="left", padx=20, pady=12)

        columns = ("ID MPL", "ФИО", "Дата рождения", "Дата сделки")
        tree2 = ttk.Treeview(w, columns=columns, show="headings")
        for col in columns: tree2.heading(col, text=col); tree2.column(col, anchor="w", width=220)
        scroll_y = ctk.CTkScrollbar(w, orientation="vertical", command=tree2.yview,
                                    button_color=CLR_ACCENT, button_hover_color=CLR_ACCENT2)
        tree2.configure(yscrollcommand=scroll_y.set)
        toast2 = ToastNotification(w); attach_context_menu(tree2, toast2)
        bf2 = ctk.CTkFrame(w, fg_color=CLR_BG); bf2.pack(pady=8)
        ctk.CTkButton(bf2, text="📤  Экспорт Excel", width=180,
                      command=lambda: export_to_excel(tree2, "Кредиты", toast2),
                      fg_color=CLR_ACCENT, hover_color=CLR_ACCENT2,
                      corner_radius=8, font=("Bahnschrift", 13)).pack()
        animate_rows(tree2, results, [""] * len(results), delay=25)
        tree2.pack(side="left", fill="both", expand=True, padx=(12, 0), pady=8)
        scroll_y.pack(side="right", fill="y", pady=8, padx=(0, 4))
    except Exception as e:
        messagebox.showerror("Ошибка", str(e))


def auto_resize(tv):
    for col in tv["columns"]:
        mw = max([len(str(tv.set(k, col))) for k in tv.get_children()] + [len(col)])
        tv.column(col, width=mw * 14)

# ════════════════════════════════════════════════════════════════════════════
#  ЗАПУСК
# ════════════════════════════════════════════════════════════════════════════
main_menu()
root.mainloop()